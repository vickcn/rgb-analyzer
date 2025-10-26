"""
彩色燈光分類驗證腳本
使用 imgData/ 目錄下的圖片驗證 HSV + 色溫 綜合分析的準確性
"""

import cv2
import numpy as np
import os
from pathlib import Path
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def rgb_to_hsv(r, g, b):
    """RGB 轉 HSV"""
    r, g, b = r / 255.0, g / 255.0, b / 255.0
    max_val = max(r, g, b)
    min_val = min(r, g, b)
    delta = max_val - min_val
    
    # 明度 (Value)
    v = max_val * 100
    
    # 飽和度 (Saturation)
    s = 0 if max_val == 0 else (delta / max_val) * 100
    
    # 色相 (Hue)
    if delta == 0:
        h = 0
    elif max_val == r:
        h = 60 * (((g - b) / delta) % 6)
    elif max_val == g:
        h = 60 * (((b - r) / delta) + 2)
    else:
        h = 60 * (((r - g) / delta) + 4)
    
    return h, s, v


def rgb_to_hsl(r, g, b):
    """RGB 轉 HSL"""
    r, g, b = r / 255.0, g / 255.0, b / 255.0
    max_val = max(r, g, b)
    min_val = min(r, g, b)
    delta = max_val - min_val
    
    # 亮度 (Lightness)
    l = (max_val + min_val) / 2
    
    # 飽和度 (Saturation)
    if delta == 0:
        s = 0
    else:
        s = delta / (1 - abs(2 * l - 1))
    
    s *= 100
    
    # 色相 (Hue)
    if delta == 0:
        h = 0
    elif max_val == r:
        h = 60 * (((g - b) / delta) % 6)
    elif max_val == g:
        h = 60 * (((b - r) / delta) + 2)
    else:
        h = 60 * (((r - g) / delta) + 4)
    
    # 如果 h 是負數，加 360
    if h < 0:
        h += 360
    
    l *= 100
    
    return h, s, l


def rgb_to_color_temp(r, g, b):
    """使用 McCamy's approximation 計算色溫"""
    # RGB 正規化
    r_norm = r / 255.0
    g_norm = g / 255.0
    b_norm = b / 255.0
    
    # Gamma 校正
    r_linear = r_norm / 12.92 if r_norm <= 0.04045 else ((r_norm + 0.055) / 1.055) ** 2.4
    g_linear = g_norm / 12.92 if g_norm <= 0.04045 else ((g_norm + 0.055) / 1.055) ** 2.4
    b_linear = b_norm / 12.92 if b_norm <= 0.04045 else ((b_norm + 0.055) / 1.055) ** 2.4
    
    # RGB → XYZ (sRGB D65)
    x = 0.4124564 * r_linear + 0.3575761 * g_linear + 0.1804375 * b_linear
    y = 0.2126729 * r_linear + 0.7151522 * g_linear + 0.0721750 * b_linear
    z = 0.0193339 * r_linear + 0.1191920 * g_linear + 0.9503041 * b_linear
    
    # 計算色度坐標
    x_xy = x / (x + y + z) if (x + y + z) != 0 else 0
    y_xy = y / (x + y + z) if (x + y + z) != 0 else 0
    
    # McCamy 公式
    n = (x_xy - 0.3320) / (0.1858 - y_xy) if (0.1858 - y_xy) != 0 else 0
    cct = 449 * n**3 + 3525 * n**2 + 6823.3 * n + 5520.33
    
    return max(0, cct)


def classify_color(hue, saturation, value, color_temp):
    """
    綜合分類色光類型
    使用 HSV + 色溫綜合分析
    針對 imgData 中的具體色光類型進行優化
    """
    # 1. 白色系列（低飽和度 + 高明度）
    if saturation < 25 and value > 75:
        if color_temp > 6500:  # 冷白光
            if 200 <= hue < 230:
                return "冰藍白"
            elif 170 <= hue < 200:
                return "白偏藍"
            elif 100 <= hue < 130:
                return "白偏綠"
            else:
                return "冰白"
        elif 3500 <= color_temp <= 4500:  # 中性白光
            if 100 <= hue < 130:
                return "白偏綠"
            elif 300 <= hue or hue < 30:
                return "白偏粉"
            else:
                return "冰暖白"
        else:  # 暖白光
            return "暖白1" if saturation > 10 else "暖白2"
    
    # 2. 香檳金（特定條件）
    if (3500 < color_temp < 4000 or color_temp < 3000) and 15 <= hue < 45 and 30 <= saturation <= 60 and value > 70:
        return "香檳金"
    
    # 3. 桃紅/粉色系列
    if (0 <= hue < 20 or hue >= 330) and saturation > 30:
        if saturation > 60:
            return "網紅A" if hue > 10 else "網紅B"
        else:
            return "桃紅" if hue > 340 or hue < 10 else "粉色"
    
    # 4. 紫色系列
    if 270 <= hue < 300:
        return "網紫" if saturation < 40 else "紫色"
    
    # 5. 藍色系列
    if 200 <= hue < 250:
        if saturation > 50:
            return "網藍A" if hue < 220 else "網藍B"
        else:
            return "網藍白"
    
    # 6. 綠色系列
    if 100 <= hue < 160:
        if saturation > 50:
            return "網草綠" if hue < 120 else "網綠"
        else:
            return "網白偏綠"
    
    # 7. 根據色相判斷通用彩色
    if 20 <= hue < 40:
        return "桃紅" if saturation > 40 else "淺紅"
    elif 40 <= hue < 60:
        return "橙色"
    elif 60 <= hue < 90:
        return "黃色"
    elif 90 <= hue < 120:
        return "黃綠色"
    elif 200 <= hue < 240:
        return "藍色"
    elif 240 <= hue < 270:
        return "紫藍色"
    elif 300 <= hue < 330:
        return "紫紅色"
    else:
        return "紅色"
    
    return "未知"


def extract_color_from_region(image, mask=None):
    """
    從圖像中提取代表顏色
    自動排除文字區域和過度邊緣
    """
    # 轉換為 HSV 以便更好地選擇顏色區域
    hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
    
    # 創建遮罩排除過暗和過亮的區域（可能是文字）
    # 過暗（V < 30）和過亮（V > 240）的區域可能是文字
    mask_valid = cv2.inRange(hsv, (0, 0, 30), (180, 255, 240))
    
    if mask is not None:
        mask_valid = cv2.bitwise_and(mask_valid, mask)
    
    # 統計有效像素
    if cv2.countNonZero(mask_valid) == 0:
        # 如果遮罩太嚴格，使用整個圖像
        mask_valid = np.ones(image.shape[:2], dtype=np.uint8) * 255
    
    # 計算所有有效像素的平均顏色
    mean_color = cv2.mean(image, mask_valid)[:3]
    
    # 轉換為 RGB
    r = int(mean_color[2])
    g = int(mean_color[1])
    b = int(mean_color[0])
    
    return r, g, b


def extract_dominant_color(image):
    """
    從圖像中提取主導顏色（採用 pysrc_v0 的方法）
    結合黑白閾值過濾和 Canny 邊緣檢測
    """
    # 轉換為 BGR (OpenCV 預設格式)
    if len(image.shape) == 3:
        if image.shape[2] == 4:  # RGBA
            bgr_image = cv2.cvtColor(image, cv2.COLOR_RGBA2BGR)
        else:
            bgr_image = image.copy()
    else:
        return extract_color_from_region(image)
    
    # ===== 方法1：使用 pysrc_v0 的邊緣檢測方法 =====
    # 1. 創建黑白遮罩（去除極端像素）
    height, width = bgr_image.shape[:2]
    mask = np.ones((height, width), dtype=np.uint8) * 255
    
    # 轉換為 RGB 進行判斷
    rgb_image = cv2.cvtColor(bgr_image, cv2.COLOR_BGR2RGB)
    
    for y in range(height):
        for x in range(width):
            pixel = rgb_image[y, x]
            r, g, b = pixel
            
            # 判斷是否為極黑色（所有通道都 < 20）
            is_extreme_black = r < 20 and g < 20 and b < 20
            
            # 判斷是否為極白色（所有通道都 > 240 且差異極小）
            is_extreme_white = (
                r > 240 and g > 240 and b > 240 and
                max(r, g, b) - min(r, g, b) < 8
            )
            
            # 只過濾極端的黑白像素，保留有色彩的區域
            if is_extreme_black or is_extreme_white:
                mask[y, x] = 0  # 設為遮蔽
    
    # 應用遮罩
    masked_image = cv2.bitwise_and(rgb_image, rgb_image, mask=mask)
    
    # 2. Canny 邊緣檢測找出色光區域
    try:
        # 轉換為灰階
        gray = cv2.cvtColor(masked_image, cv2.COLOR_RGB2GRAY)
        
        # 高斯模糊減少噪音
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        
        # Canny 邊緣檢測
        edges = cv2.Canny(blurred, 50, 150)
        
        # 形態學操作連接邊緣
        kernel = np.ones((3, 3), np.uint8)
        edges = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, kernel)
        
        # 尋找輪廓
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        # 過濾小輪廓，並且檢查是否為文字框
        min_area = 100
        valid_contours = []
        
        print(f"Canny 檢測到 {len(contours)} 個輪廓")
        
        for i, cnt in enumerate(contours):
            area = cv2.contourArea(cnt)
            if area < min_area:
                continue
            
            # 檢查輪廓的平均顏色
            mask_temp = np.zeros((height, width), dtype=np.uint8)
            cv2.drawContours(mask_temp, [cnt], -1, 255, -1)
            
            # 計算這個區域的平均顏色
            region_pixels = rgb_image[mask_temp > 0]
            if len(region_pixels) > 0:
                avg_color = np.mean(region_pixels, axis=0)
                r_avg, g_avg, b_avg = avg_color
                
                # 排除文字框區域（更精確的判定）
                # 1. 排除極端純白區域（文字框背景）
                if (r_avg > 245 and g_avg > 245 and b_avg > 245 and 
                    max(r_avg, g_avg, b_avg) - min(r_avg, g_avg, b_avg) < 8):
                    print(f"跳過純白區域: RGB({r_avg:.0f}, {g_avg:.0f}, {b_avg:.0f})")
                    continue  # 跳過純白文字框
                
                # 2. 排除偏正白的大面積區域（可能是文字框）
                avg_brightness = (r_avg + g_avg + b_avg) / 3
                color_deviation = max(r_avg, g_avg, b_avg) - min(r_avg, g_avg, b_avg)
                
                # 整塊偏正白：亮度很高 + 色彩差異很小 + 面積較大
                is_whitish_block = (
                    avg_brightness > 220 and  # 很亮
                    color_deviation < 15 and  # 色彩差異很小（接近白色）
                    area > 1000  # 面積較大（整塊區域）
                )
                
                if is_whitish_block:
                    print(f"跳過偏正白大區域: RGB({r_avg:.0f}, {g_avg:.0f}, {b_avg:.0f}), 面積={area:.0f}, 色差={color_deviation:.1f}")
                    continue  # 跳過偏正白文字框
                
                # 3. 排除純黑區域
                if r_avg < 10 and g_avg < 10 and b_avg < 10:
                    print(f"跳過純黑區域: RGB({r_avg:.0f}, {g_avg:.0f}, {b_avg:.0f})")
                    continue  # 跳過純黑區域
                
                # 新增：檢查區域形狀（文字框通常是矩形且長寬比接近1）
                x, y, w, h = cv2.boundingRect(cnt)
                aspect_ratio = max(w, h) / max(min(w, h), 1)  # 長寬比
                extent = area / (w * h)  # 區域面積與邊界框面積比
                
                # 檢查是否為文字框特徵（更嚴格的條件）
                # 文字框通常：長寬比合理、extent 很小、且飽和度極低
                max_val = max(r_avg, g_avg, b_avg)
                min_val = min(r_avg, g_avg, b_avg)
                saturation_temp = max_val - min_val
                
                is_text_box = (
                    aspect_ratio < 3 and  # 不是極細長條
                    extent < 0.3 and  # 很稀疏（文字通常佔很小空間）
                    saturation_temp < 15 and  # 飽和度極低（接近灰色）
                    50 < (r_avg + g_avg + b_avg) / 3 < 200  # 中等亮度
                )
                
                if is_text_box:
                    print(f"跳過疑似文字框區域: 面積={area:.0f}, 長寬比={aspect_ratio:.2f}, extent={extent:.2f}, 飽和度={saturation_temp:.1f}")
                    continue  # 跳過疑似文字框
                
                # 檢查是否為中性灰色（可能是邊框或背景）
                max_val = max(r_avg, g_avg, b_avg)
                min_val = min(r_avg, g_avg, b_avg)
                avg_brightness = (r_avg + g_avg + b_avg) / 3
                
                # 只排除非常低飽和度的中性灰色區域
                # 保留「白偏X」類型的顏色（有輕微色偏但整體偏亮）
                is_neutral_gray = (
                    max_val - min_val < 8 and  # 極低色彩差異
                    80 < avg_brightness < 180 and  # 中等亮度
                    area < 300  # 且面積不大（可能是雜訊）
                )
                
                if is_neutral_gray:
                    continue  # 跳過中性灰色小區域
                
                # 檢查區域大小，優先保留較大的區域
                if area > 500:  # 大區域直接保留
                    valid_contours.append(cnt)
                    continue
            
            valid_contours.append(cnt)
        
        if len(valid_contours) > 0:
            # 根據區域面積和色彩信息選擇最合適的區域
            contour_info = []
            
            for cnt in valid_contours:
                area = cv2.contourArea(cnt)
                mask_temp = np.zeros((height, width), dtype=np.uint8)
                cv2.drawContours(mask_temp, [cnt], -1, 255, -1)
                
                region_pixels = rgb_image[mask_temp > 0]
                if len(region_pixels) > 0:
                    avg_color = np.mean(region_pixels, axis=0)
                    r_avg, g_avg, b_avg = avg_color
                    
                    # 計算飽和度和亮度
                    max_val = max(r_avg, g_avg, b_avg)
                    min_val = min(r_avg, g_avg, b_avg)
                    saturation = max_val - min_val
                    brightness = (r_avg + g_avg + b_avg) / 3
                    
                    # 計算權重（面積 + 飽和度 + 亮度）
                    # 飽和度高的區域優先（更可能是有色光）
                    # 亮度適中的區域優先（太亮可能是背景或文字框）
                    brightness_weight = 1 if 50 < brightness < 230 else 0.5
                    
                    # 調整：低飽和度區域懲罰（放寬條件）
                    # 只對極低飽和度區域懲罰，保留有色光
                    if saturation < 10:
                        saturation_penalty = 0.2  # 嚴重懲罰（極低飽和度）
                    elif saturation < 25:
                        saturation_penalty = 0.6  # 輕微懲罰（低飽和度）
                    else:
                        saturation_penalty = 1.0  # 無懲罰（正常飽和度）
                    
                    score = area * (1 + saturation / 255.0) * brightness_weight * saturation_penalty
                    
                    contour_info.append({
                        'contour': cnt,
                        'area': area,
                        'rgb': (int(r_avg), int(g_avg), int(b_avg)),
                        'saturation': saturation,
                        'brightness': brightness,
                        'score': score
                    })
            
            # 按分數排序，選擇最合適的區域
            contour_info.sort(key=lambda x: x['score'], reverse=True)
            
            if len(contour_info) > 0:
                best_region = contour_info[0]
                print(f"選擇最佳區域:")
                print(f"  面積: {best_region['area']:.0f} 像素")
                print(f"  RGB: {best_region['rgb']}")
                print(f"  飽和度: {best_region['saturation']:.1f}")
                print(f"  亮度: {best_region['brightness']:.1f}")
                print(f"  分數: {best_region['score']:.1f}")
                
                # 使用最佳區域
                region_mask = np.zeros((height, width), dtype=np.uint8)
                cv2.drawContours(region_mask, [best_region['contour']], -1, 255, -1)
                
                # 統計這個區域的像素
                valid_pixels = rgb_image[region_mask > 0]
                
                if len(valid_pixels) > 0:
                    avg_rgb = np.mean(valid_pixels, axis=0)
                    return (int(avg_rgb[0]), int(avg_rgb[1]), int(avg_rgb[2]))
    except Exception as e:
        print(f"邊緣檢測失敗: {e}")
    
    # ===== 方法2：回退到簡單的平均值（只在有效遮罩區域） =====
    # 只統計非黑白像素
    valid_pixels = rgb_image[mask > 0]
    
    if len(valid_pixels) > 0:
        avg_rgb = np.mean(valid_pixels, axis=0)
        return (int(avg_rgb[0]), int(avg_rgb[1]), int(avg_rgb[2]))
    
    # 最後回退：整個圖像的平均值
    return extract_color_from_region(image)


def draw_edge_detection(image):
    """
    繪製邊緣檢測結果（用於視覺化）
    參考 pysrc_v0 的邊緣框繪製方法
    
    Returns:
        edge_image: 標註了檢測區域的圖片
    """
    # 1. 創建黑白遮罩
    height, width = image.shape[:2]
    mask = np.ones((height, width), dtype=np.uint8) * 255
    
    # 轉換為 RGB 進行判斷
    rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    
    for y in range(height):
        for x in range(width):
            pixel = rgb_image[y, x]
            r, g, b = pixel
            
            # 判斷是否為黑色或白色
            is_black = r < 30 and g < 30 and b < 30
            is_white = r > 225 and g > 225 and b > 225
            
            if is_black or is_white:
                mask[y, x] = 0
    
    # 2. Canny 邊緣檢測
    edge_image = image.copy()
    
    try:
        # 應用遮罩
        masked_image = cv2.bitwise_and(rgb_image, rgb_image, mask=mask)
        
        # 轉換為灰階
        gray = cv2.cvtColor(masked_image, cv2.COLOR_RGB2GRAY)
        
        # 高斯模糊
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)
        
        # Canny 邊緣檢測
        edges = cv2.Canny(blurred, 50, 150)
        
        # 形態學操作
        kernel = np.ones((3, 3), np.uint8)
        edges = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, kernel)
        
        # 尋找輪廓
        contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        # 使用與主要演算法相同的過濾邏輯
        min_area = 100
        valid_contours = []
        
        print(f"找到 {len(contours)} 個輪廓")
        
        for i, cnt in enumerate(contours):
            area = cv2.contourArea(cnt)
            if area < min_area:
                continue
            
            # 檢查輪廓的平均顏色（與主演算法一致）
            mask_temp = np.zeros((height, width), dtype=np.uint8)
            cv2.drawContours(mask_temp, [cnt], -1, 255, -1)
            
            region_pixels = rgb_image[mask_temp > 0]
            if len(region_pixels) > 0:
                avg_color = np.mean(region_pixels, axis=0)
                r_avg, g_avg, b_avg = avg_color
                
                # 使用相同的過濾條件
                # 1. 排除極端純白區域
                if (r_avg > 245 and g_avg > 245 and b_avg > 245 and 
                    max(r_avg, g_avg, b_avg) - min(r_avg, g_avg, b_avg) < 8):
                    continue
                
                # 2. 排除偏正白的大面積區域
                avg_brightness = (r_avg + g_avg + b_avg) / 3
                color_deviation = max(r_avg, g_avg, b_avg) - min(r_avg, g_avg, b_avg)
                
                is_whitish_block = (
                    avg_brightness > 220 and
                    color_deviation < 15 and
                    area > 1000
                )
                
                if is_whitish_block:
                    continue
                
                # 3. 排除純黑區域
                if r_avg < 10 and g_avg < 10 and b_avg < 10:
                    continue
                
                # 檢查是否為中性灰色（可能是邊框或背景）
                # 使用已計算的 color_deviation 和 avg_brightness
                is_neutral_gray = (
                    color_deviation < 8 and
                    80 < avg_brightness < 180 and
                    area < 300
                )
                
                if is_neutral_gray:
                    continue
                
                if area > 500:
                    valid_contours.append(cnt)
                    continue
            
            valid_contours.append(cnt)
        
        # 如果有多個區域，計算分數並排序
        if len(valid_contours) > 1:
            contour_info = []
            for cnt in valid_contours:
                area = cv2.contourArea(cnt)
                mask_temp = np.zeros((height, width), dtype=np.uint8)
                cv2.drawContours(mask_temp, [cnt], -1, 255, -1)
                
                region_pixels = rgb_image[mask_temp > 0]
                if len(region_pixels) > 0:
                    avg_color = np.mean(region_pixels, axis=0)
                    r_avg, g_avg, b_avg = avg_color
                    max_val = max(r_avg, g_avg, b_avg)
                    min_val = min(r_avg, g_avg, b_avg)
                    saturation = max_val - min_val
                    brightness = (r_avg + g_avg + b_avg) / 3
                    
                    brightness_weight = 1 if 50 < brightness < 230 else 0.5
                    
                    if saturation < 10:
                        saturation_penalty = 0.2
                    elif saturation < 25:
                        saturation_penalty = 0.6
                    else:
                        saturation_penalty = 1.0
                    
                    score = area * (1 + saturation / 255.0) * brightness_weight * saturation_penalty
                    
                    contour_info.append({
                        'contour': cnt,
                        'area': area,
                        'score': score
                    })
            
            contour_info.sort(key=lambda x: x['score'], reverse=True)
            valid_contours = [info['contour'] for info in contour_info]
        
        # 繪製輪廓
        cv2.drawContours(edge_image, valid_contours, -1, (0, 255, 0), 2)
        
        # 繪製矩形框和編號（最佳區域用紅色，其他用藍色）
        for i, contour in enumerate(valid_contours):
            x, y, w, h = cv2.boundingRect(contour)
            area = cv2.contourArea(contour)
            
            # 最佳區域用紅色框，其他用藍色框
            color = (0, 0, 255) if i == 0 else (255, 0, 0)  # BGR格式
            cv2.rectangle(edge_image, (x, y), (x + w, y + h), color, 2)
            
            # 添加編號和面積
            label = f"{i + 1} ({area:.0f})"
            cv2.putText(edge_image, label, (x, y - 10),
                       cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)
    except Exception as e:
        print(f"繪製邊緣框失敗: {e}")
    
    return edge_image


def analyze_image(image_path):
    """
    分析單張圖片
    
    流程：
    1. 讀取圖片
    2. 使用黑白閾值過濾（R<30, G<30, B<30 或 R>225, G>225, B>225）
    3. Canny 邊緣檢測找出色光區域
    4. 過濾小面積區域（面積 > 100 像素）
    5. 計算有效區域的 RGB 平均值
    6. 轉換為 HSV 和色溫
    7. 分類色光類型
    """
    print(f"\n{'='*60}")
    print(f"分析圖片: {os.path.basename(image_path)}")
    print(f"{'='*60}")
    
    # 讀取圖片
    img = cv2.imread(image_path)
    if img is None:
        print(f"⚠️  無法讀取圖片: {image_path}")
        return None
    
    print(f"圖片尺寸: {img.shape[1]} x {img.shape[0]}")
    
    # 使用 pysrc_v0 的方法提取主導顏色
    # 流程：黑白過濾 → Canny 邊緣檢測 → 區域提取 → RGB 計算
    r, g, b = extract_dominant_color(img)
    print(f"\n主導顏色: RGB({r}, {g}, {b})")
    print(f"（已去除黑白干擾，專注色光區域）")
    
    # 計算 HSV
    hue, sat, val = rgb_to_hsv(r, g, b)
    print(f"HSV: H={hue:.1f}°, S={sat:.1f}%, V={val:.1f}%")
    
    # 計算 HSL
    hsl_h, hsl_s, hsl_l = rgb_to_hsl(r, g, b)
    print(f"HSL: H={hsl_h:.1f}°, S={hsl_s:.1f}%, L={hsl_l:.1f}%")
    
    # 計算色溫
    color_temp = rgb_to_color_temp(r, g, b)
    print(f"色溫: {color_temp:.0f}K")
    
    # 分類
    classification = classify_color(hue, sat, val, color_temp)
    print(f"分類結果: {classification}")
    
    return {
        'filename': os.path.basename(image_path),
        'rgb': (r, g, b),
        'hsv': (hue, sat, val),
        'hsl': (hsl_h, hsl_s, hsl_l),
        'color_temp': color_temp,
        'classification': classification
    }


def main():
    """主程式"""
    # 圖片目錄
    img_dir = Path(__file__).parent.parent / "imgData"
    
    if not img_dir.exists():
        print(f"❌ 目錄不存在: {img_dir}")
        return
    
    # 獲取所有 PNG 圖片
    png_files = sorted(img_dir.glob("*.png"))
    
    if not png_files:
        print(f"❌ 未找到 PNG 圖片在: {img_dir}")
        return
    
    print(f"找到 {len(png_files)} 張圖片")
    
    # 分析結果
    results = []
    expected_classes = []
    
    for png_file in png_files:
        # 從檔案名提取預期分類
        filename = png_file.stem
        expected_class = filename  # 檔案名本身就是分類
        
        expected_classes.append(expected_class)
        
        # 分析圖片
        result = analyze_image(str(png_file))
        if result:
            results.append(result)
    
    # 統計和驗證
    print(f"\n\n{'='*60}")
    print("分析結果總結")
    print(f"{'='*60}\n")
    
    correct = 0
    total = len(results)
    
    for i, (result, expected) in enumerate(zip(results, expected_classes)):
        filename = result['filename']
        predicted = result['classification']
        
        print(f"{filename:25s} | 預測: {predicted:12s}", end="")
        
        # 簡化比對：只比較主要特徵
        # 例如「網白偏綠」和「白偏綠」視為同一類
        expected_clean = expected.replace('網', '').strip()
        predicted_clean = predicted.replace('網', '').strip()
        
        # 特殊處理「暖白1」和「暖白2」都算暖白
        if '暖白' in expected_clean and '暖白' in predicted_clean:
            match = True
        elif predicted_clean == expected_clean or expected_clean in predicted_clean or predicted_clean in expected_clean:
            match = True
        else:
            match = False
        
        if match:
            print(f" | ✓ 匹配 ({expected})")
            correct += 1
        else:
            print(f" | ✗ 不匹配 (預期: {expected})")
    
    print(f"\n{'='*60}")
    if total > 0:
        accuracy = (correct / total) * 100
        print(f"準確率: {correct}/{total} ({accuracy:.1f}%)")
    print(f"{'='*60}")
    
    # 匯出詳細結果和邊緣框圖片
    output_dir = Path(__file__).parent / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    edge_dir = output_dir / "邊緣框檔"
    edge_dir.mkdir(parents=True, exist_ok=True)
    
    print(f"\n生成邊緣框圖片中...")
    
    # 為每張圖生成邊緣框圖片
    for i, png_file in enumerate(png_files):
        if i < len(results):
            try:
                img = cv2.imread(str(png_file))
                if img is not None:
                    # 使用相同的提取邏輯生成邊緣框
                    edge_image = draw_edge_detection(img)
                    
                    # 儲存邊緣框圖片
                    output_edge_file = edge_dir / f"{png_file.stem}_邊緣框.png"
                    cv2.imwrite(str(output_edge_file), edge_image)
                    print(f"✓ 已生成: {output_edge_file.name}")
            except Exception as e:
                print(f"⚠️  生成 {png_file.name} 的邊緣框失敗: {e}")
    
    # 匯出詳細文字結果
    output_file = output_dir / "color_analysis_result.txt"
    
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("彩色燈光分析結果\n")
        f.write("="*60 + "\n\n")
        
        for result in results:
            f.write(f"檔案: {result['filename']}\n")
            f.write(f"RGB: {result['rgb']}\n")
            f.write(f"HSV: H={result['hsv'][0]:.1f}°, S={result['hsv'][1]:.1f}%, V={result['hsv'][2]:.1f}%\n")
            f.write(f"色溫: {result['color_temp']:.0f}K\n")
            f.write(f"分類: {result['classification']}\n")
            f.write("-"*60 + "\n")
    
    print(f"\n詳細結果已儲存至:")
    print(f"  文字報告: {output_file}")
    print(f"  邊緣框檔: {edge_dir}")
    
    # 匯出 Excel 報告
    excel_file = output_dir / "color_analysis_result.xlsx"
    export_to_excel(results, excel_file)
    print(f"  Excel 報告: {excel_file}")


def export_to_excel(results, output_file):
    """
    將分析結果匯出到 Excel
    包含圖片路徑、演算參數、RGB、HSV、色溫、分類結果
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "彩色燈光分析報告"
    
    # 設定標題
    headers = [
        "編號", "圖片名稱", "RGB", "R", "G", "B",
        "HSV", "H (色相)", "S (飽和度)", "V (明度)",
        "HSL", "HSL_H", "HSL_S", "HSL_L",
        "色溫 (K)", "色溫描述", "分類結果", "邊緣框檔"
    ]
    
    ws.append(headers)
    
    # 設定標題行樣式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 填充數據
    for idx, result in enumerate(results, 1):
        row = [
            idx,
            result['filename'],
            f"RGB({result['rgb'][0]}, {result['rgb'][1]}, {result['rgb'][2]})",
            result['rgb'][0],
            result['rgb'][1],
            result['rgb'][2],
            f"HSV({result['hsv'][0]:.1f}°, {result['hsv'][1]:.1f}%, {result['hsv'][2]:.1f}%)",
            f"{result['hsv'][0]:.1f}°",
            f"{result['hsv'][1]:.1f}%",
            f"{result['hsv'][2]:.1f}%",
            f"HSL({result['hsl'][0]:.1f}°, {result['hsl'][1]:.1f}%, {result['hsl'][2]:.1f}%)",
            f"{result['hsl'][0]:.1f}°",
            f"{result['hsl'][1]:.1f}%",
            f"{result['hsl'][2]:.1f}%",
            f"{result['color_temp']:.0f}",
            get_color_temp_description(result['color_temp']),
            result['classification'],
            f"{result['filename'].rsplit('.', 1)[0]}_邊緣框.png"
        ]
        ws.append(row)
    
    # 自動調整列寬
    for col_num, header in enumerate(headers, 1):
        max_length = len(header)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            if row[0].value:
                max_length = max(max_length, len(str(row[0].value)))
        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2
    
    # 凍結首行
    ws.freeze_panes = "A2"
    
    # 儲存
    wb.save(output_file)
    print(f"✓ Excel 報告已生成: {output_file.name}")


def get_color_temp_description(cct):
    """根據色溫返回描述"""
    if cct < 2000:
        return "極暖光（燭光）"
    elif cct < 3000:
        return "暖光（鎢絲燈）"
    elif cct < 3500:
        return "暖白光"
    elif cct < 4500:
        return "中性白光"
    elif cct < 5500:
        return "自然光"
    elif cct < 6500:
        return "日光"
    elif cct < 8000:
        return "冷白光"
    elif cct < 10000:
        return "冷光（陰天）"
    else:
        return "極冷光（藍天）"


if __name__ == "__main__":
    main()

